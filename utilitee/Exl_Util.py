import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

class EXL:

    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    def getCoulmnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_coulmn)

    def readData(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(rownum, columnno).value

    def writeData(file, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(rownum, columnno).value = data
        workbook.save(file)

    def including_data(self, xl_file, sheet1):
        add = []
        first_name = EXL.readData(xl_file, sheet1, 4, 4)
        mid_name = EXL.readData(xl_file, sheet1, 4, 5)
        last_name = EXL.readData(xl_file, sheet1, 4, 6)
        empl_id = EXL.readData(xl_file, sheet1, 4, 7)
        other_id = EXL.readData(xl_file, sheet1, 4, 8)
        driving_no = EXL.readData(xl_file, sheet1, 4, 9)
        lice_ex_dt = EXL.readData(xl_file, sheet1, 4, 10)
        date_ob = EXL.readData(xl_file, sheet1, 4, 11)
        gender = EXL.readData(xl_file, sheet1, 4, 12)
        test_field = EXL.readData(xl_file, sheet1, 4, 13)

        ad = (first_name, mid_name, last_name, empl_id,other_id, driving_no, lice_ex_dt, date_ob, gender, test_field)
        ad.append(add)
        return add

    def editing_datas(self, exl_file_path, sheet2):
        edit = []
        search_id= EXL.readData(exl_file_path, sheet2, 5, 7)
        last_name = EXL.readData(exl_file_path, sheet2, 5, 6)
        empo_id = EXL.readData(exl_file_path, sheet2, 5, 7)

        redact = (search_id, last_name, empo_id)
        redact.append(edit)
        return edit
